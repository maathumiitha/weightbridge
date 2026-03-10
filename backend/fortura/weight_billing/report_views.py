# report_views.py
from rest_framework import viewsets
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Sum, Q
from django.http import HttpResponse
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from .models import WeightRecord
from .serializers import (
    ReportFilterSerializer, AggregatedReportSerializer,
    WeightRecordWithDropsSerializer
)
from .utils import create_audit_log

# ==================== SECURITY IMPORTS ====================
from .models_security_data_management import log_security_action


class ReportViewSet(viewsets.ViewSet):
    """ViewSet for generating reports with security audit logging"""
    permission_classes = [IsAuthenticated]

    @staticmethod
    def _to_float(value, default=0.0):
        """Safely convert Decimal/None values to float for export rendering."""
        if value is None:
            return default
        try:
            return float(value)
        except (TypeError, ValueError):
            return default
    
    def _build_filters(self, data):
        """Helper to build query filters"""
        filters = Q()
        
        # Exclude soft-deleted records
        filters &= Q(is_deleted=False)
        
        if 'start_date' in data:
            filters &= Q(date__gte=data['start_date'])
        if 'end_date' in data:
            filters &= Q(date__lte=data['end_date'])
        if 'shift' in data:
            filters &= Q(shift=data['shift'])
        if 'vehicle' in data:
            filters &= Q(vehicle_id=data['vehicle'])
        if 'customer' in data:
            filters &= Q(customer_id=data['customer'])
        if 'operator' in data:
            filters &= (
                Q(operator_first_weight_id=data['operator']) |
                Q(operator_second_weight_id=data['operator'])
            )
        return filters
    
    @action(detail=False, methods=['post'])
    def fetch_records(self, request):
        """Fetch records based on filters"""
        serializer = ReportFilterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        filters = self._build_filters(serializer.validated_data)
        records = WeightRecord.objects.filter(filters).select_related(
            'customer', 'operator_first_weight', 'operator_second_weight', 'vehicle'
        ).prefetch_related('drops').order_by('-date', 'shift')
        
        # ==================== SECURITY: AUDIT LOG ====================
        log_security_action(
            action='VIEW',
            user=request.user,
            affected_model='WeightRecord',
            notes=f'Fetched {records.count()} weight records for report',
            request=request
        )
        
        serializer = WeightRecordWithDropsSerializer(records, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def aggregate_totals(self, request):
        """Aggregate totals based on filters"""
        serializer = ReportFilterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        filters = self._build_filters(serializer.validated_data)
        records = WeightRecord.objects.filter(filters).select_related(
            'customer', 'operator_first_weight', 'operator_second_weight', 'vehicle'
        )
        
        aggregates = records.aggregate(
            total_gross_weight=Sum('gross_weight'),
            total_tare_weight=Sum('tare_weight'),
            total_net_weight=Sum('net_weight'),
            total_amount=Sum('total_amount')
        )
        
        # Prepare detailed records
        detailed_records = []
        for record in records:
            detailed_records.append({
                'slip_number': record.slip_number,
                'date': record.date,
                'shift': record.shift,
                'vehicle_number': record.vehicle.vehicle_number,
                'customer_driver_name': record.customer.driver_name,
                'operator_first_weight_name': record.operator_first_weight.employee_name if record.operator_first_weight else '',
                'operator_second_weight_name': record.operator_second_weight.employee_name if record.operator_second_weight else '',
                'first_weight': record.first_weight or 0,
                'first_weight_time': record.first_weight_time or record.created_at,
                'second_weight': record.second_weight,
                'second_weight_time': record.second_weight_time,
                'gross_weight': record.gross_weight or 0,
                'tare_weight': record.tare_weight or 0,
                'net_weight': record.net_weight or 0,
                'material_type': record.material_type or '',
                'rate_per_unit': record.rate_per_unit or 0,
                'total_amount': record.total_amount or 0,
                'status': record.status,
                'remarks': record.remarks or ''
            })
        
        response_data = {
            'total_records': records.count(),
            'total_gross_weight': aggregates['total_gross_weight'] or 0,
            'total_tare_weight': aggregates['total_tare_weight'] or 0,
            'total_net_weight': aggregates['total_net_weight'] or 0,
            'total_amount': aggregates['total_amount'] or 0,
            'records': detailed_records
        }
        
        # ==================== SECURITY: AUDIT LOG ====================
        log_security_action(
            action='VIEW',
            user=request.user,
            affected_model='WeightRecord',
            new_values={
                'total_records': records.count(),
                'total_amount': str(aggregates['total_amount'] or 0)
            },
            notes=f'Generated aggregate report with {records.count()} records, total amount: {aggregates["total_amount"] or 0}',
            request=request
        )
        
        serializer = AggregatedReportSerializer(data=response_data)
        serializer.is_valid(raise_exception=True)
        
        return Response(serializer.data)
    
    @action(detail=False, methods=['post'])
    def export_excel(self, request):
        """Export report to Excel"""
        serializer = ReportFilterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        filters = self._build_filters(serializer.validated_data)
        records = WeightRecord.objects.filter(filters).select_related(
            'customer', 'operator_first_weight', 'operator_second_weight', 'vehicle'
        ).order_by('-date', 'shift')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Weight Report"
        
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Headers
        headers = [
            'Date', 'Shift', 'Vehicle No', 'Driver Name', 'Operator (1st)', 'Operator (2nd)',
            'Gross Weight', 'Tare Weight', 'Net Weight',
            'Material', 'Rate/Unit', 'Total Amount', 'Multi-Drop', 'Remarks'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row, record in enumerate(records, 2):
            ws.cell(row=row, column=1, value=record.date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=2, value=record.shift)
            ws.cell(row=row, column=3, value=record.vehicle.vehicle_number)
            ws.cell(row=row, column=4, value=record.customer.driver_name)
            ws.cell(row=row, column=5, value=record.operator_first_weight.employee_name if record.operator_first_weight else '')
            ws.cell(row=row, column=6, value=record.operator_second_weight.employee_name if record.operator_second_weight else '')
            ws.cell(row=row, column=7, value=self._to_float(record.gross_weight))
            ws.cell(row=row, column=8, value=self._to_float(record.tare_weight))
            ws.cell(row=row, column=9, value=self._to_float(record.net_weight))
            ws.cell(row=row, column=10, value=record.material_type or '')
            ws.cell(row=row, column=11, value=self._to_float(record.rate_per_unit))
            ws.cell(row=row, column=12, value=self._to_float(record.total_amount))
            ws.cell(row=row, column=13, value='Yes' if record.is_multi_drop else 'No')
            ws.cell(row=row, column=14, value=record.remarks or '')
        
        # Add totals
        aggregates = records.aggregate(
            total_gross=Sum('gross_weight'),
            total_tare=Sum('tare_weight'),
            total_net=Sum('net_weight'),
            total_amount=Sum('total_amount')
        )
        
        total_row = len(records) + 2
        ws.cell(row=total_row, column=6, value="TOTALS:").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=float(aggregates['total_gross'] or 0)).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=float(aggregates['total_tare'] or 0)).font = Font(bold=True)
        ws.cell(row=total_row, column=9, value=float(aggregates['total_net'] or 0)).font = Font(bold=True)
        ws.cell(row=total_row, column=12, value=float(aggregates['total_amount'] or 0)).font = Font(bold=True)
        
        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            ws.column_dimensions[column].width = max_length + 2
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create regular audit log
        create_audit_log(
            weight_record=None,
            action='EXPORT',
            request=request,
            notes=f'Exported {records.count()} records to Excel'
        )
        
        # ==================== SECURITY: AUDIT LOG ====================
        filename = f'weight_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        log_security_action(
            action='VIEW',
            user=request.user,
            affected_model='WeightRecord',
            new_values={
                'export_type': 'EXCEL',
                'record_count': records.count(),
                'filename': filename,
                'total_amount': str(aggregates['total_amount'] or 0)
            },
            notes=f'Exported {records.count()} records to Excel: {filename}',
            request=request
        )
        
        # Create response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
    
    @action(detail=False, methods=['post'])
    def export_pdf(self, request):
        """Export report to PDF"""
        serializer = ReportFilterSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        
        filters = self._build_filters(serializer.validated_data)
        records = WeightRecord.objects.filter(filters).select_related(
            'customer', 'operator_first_weight', 'operator_second_weight', 'vehicle'
        ).order_by('-date', 'shift')
        
        # Create PDF
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
        elements = []
        
        styles = getSampleStyleSheet()
        
        # Title
        title = Paragraph("<b>Weight Billing Report</b>", styles['Title'])
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Table data
        table_data = [[
            'Date', 'Shift', 'Vehicle', 'Driver', 'Operator(1st)', 'Operator(2nd)',
            'Gross Wt', 'Tare Wt', 'Net Wt', 'Material', 'Rate', 'Amount'
        ]]
        
        for record in records:
            table_data.append([
                record.date.strftime('%Y-%m-%d'),
                record.shift,
                record.vehicle.vehicle_number,
                record.customer.driver_name,
                record.operator_first_weight.employee_name if record.operator_first_weight else '',
                record.operator_second_weight.employee_name if record.operator_second_weight else '',
                f"{self._to_float(record.gross_weight):.2f}",
                f"{self._to_float(record.tare_weight):.2f}",
                f"{self._to_float(record.net_weight):.2f}",
                record.material_type or '',
                f"{self._to_float(record.rate_per_unit):.2f}",
                f"{self._to_float(record.total_amount):.2f}"
            ])
        
        # Add totals
        aggregates = records.aggregate(
            total_gross=Sum('gross_weight'),
            total_tare=Sum('tare_weight'),
            total_net=Sum('net_weight'),
            total_amount=Sum('total_amount')
        )
        
        table_data.append([
            '', '', '', '', '', 'TOTALS:',
            f"{aggregates['total_gross'] or 0:.2f}",
            f"{aggregates['total_tare'] or 0:.2f}",
            f"{aggregates['total_net'] or 0:.2f}",
            '', '',
            f"{aggregates['total_amount'] or 0:.2f}"
        ])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'), 
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, -1), (-1, -1), colors.beige),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(table)
        
        # Build PDF
        doc.build(elements)
        
        # Get PDF value
        pdf = buffer.getvalue()
        buffer.close()
        
        # Create regular audit log
        create_audit_log(
            weight_record=None,
            action='EXPORT',
            request=request,
            notes=f'Exported {records.count()} records to PDF'
        )
        
        # ==================== SECURITY: AUDIT LOG ====================
        filename = f'weight_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        log_security_action(
            action='VIEW',
            user=request.user,
            affected_model='WeightRecord',
            new_values={
                'export_type': 'PDF',
                'record_count': records.count(),
                'filename': filename,
                'total_amount': str(aggregates['total_amount'] or 0)
            },
            notes=f'Exported {records.count()} records to PDF: {filename}',
            request=request
        )
         
        # Create response
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename={filename}'
        response.write(pdf)
         
        return response
